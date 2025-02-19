#!/usr/bin/env python3
import openpyxl
from mcp.server.models import InitializationOptions
import mcp.types as types
from mcp.server import NotificationOptions, Server
import mcp.server.stdio
import os
import json

server = Server("excel-reader-server")

@server.list_tools()
async def handle_list_tools() -> list[types.Tool]:
    """
    List available tools.
    Each tool specifies its arguments using JSON Schema validation.
    """
    return [
        types.Tool(
            name="read_excel",
            description="Read content from Excel (xlsx) files",
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path to the Excel file"
                    }
                },
                "required": ["file_path"]
            }
        ),
        types.Tool(
            name="read_excel_by_sheet_name",
            description="Read content from a specific sheet by name in Excel (xlsx) files. Reads first sheet if sheet_name not provided.",
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path to the Excel file"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Name of the sheet to read (optional, defaults to first sheet)"
                    }
                },
                "required": ["file_path"]
            }
        ),
        types.Tool(
            name="read_excel_by_sheet_index",
            description="Read content from a specific sheet by index in Excel (xlsx) files. Reads first sheet (index 0) if sheet_index not provided.",
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path to the Excel file"
                    },
                    "sheet_index": {
                        "type": "integer",
                        "description": "Index of the sheet to read (optional, defaults to 0)",
                        "minimum": 0
                    }
                },
                "required": ["file_path"]
            }
        )
    ]

@server.call_tool()
async def handle_call_tool(
    name: str, arguments: dict | None
) -> list[types.TextContent | types.ImageContent | types.EmbeddedResource]:    
    """
    Handle tool execution requests.
    """
    if name not in ["read_excel", "read_excel_by_sheet_name", "read_excel_by_sheet_index"]:
        raise ValueError(f"Unknown tool: {name}")
    
    if not arguments:
        raise ValueError("Missing arguments")

    file_path = arguments.get("file_path")
    if not file_path:
        raise ValueError("file_path is required")

    abs_path = os.path.abspath(file_path)
    if not os.path.exists(abs_path):
        raise ValueError(f"Excel file not found: {file_path}")
    
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(abs_path, data_only=True)
        result = {}
        
        if name == "read_excel":
            # Original functionality - process all sheets
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = []
                
                for row in sheet.rows:
                    row_data = [str(cell.value) if cell.value is not None else "" for cell in row]
                    sheet_data.append(row_data)
                
                result[sheet_name] = sheet_data
        
        elif name == "read_excel_by_sheet_name":
            # Get sheet by name, default to first sheet if not specified
            sheet_name = arguments.get("sheet_name")
            if not sheet_name:
                sheet_name = workbook.sheetnames[0]
            elif sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
                
            sheet = workbook[sheet_name]
            sheet_data = []
            
            for row in sheet.rows:
                row_data = [str(cell.value) if cell.value is not None else "" for cell in row]
                sheet_data.append(row_data)
            
            result[sheet_name] = sheet_data
            
        elif name == "read_excel_by_sheet_index":
            # Get sheet by index, default to 0 if not specified
            sheet_index = arguments.get("sheet_index", 0)
            if sheet_index < 0 or sheet_index >= len(workbook.sheetnames):
                raise ValueError(f"Sheet index {sheet_index} is out of range")
                
            sheet_name = workbook.sheetnames[sheet_index]
            sheet = workbook[sheet_name]
            sheet_data = []
            
            for row in sheet.rows:
                row_data = [str(cell.value) if cell.value is not None else "" for cell in row]
                sheet_data.append(row_data)
            
            result[sheet_name] = sheet_data

        return [
            types.TextContent(
                type="text",
                text=json.dumps(result, indent=2)
            )
        ]
        
    except Exception as e:
        raise ValueError(f"Error reading Excel file: {str(e)}")

async def main():
    # Run the server using stdin/stdout streams
    async with mcp.server.stdio.stdio_server() as (read_stream, write_stream):
        await server.run(
            read_stream,
            write_stream,
            InitializationOptions(
                server_name="excel-reader-server",
                server_version="0.1.0",
                capabilities=server.get_capabilities(
                    notification_options=NotificationOptions(),
                    experimental_capabilities={},
                ),
            ),
        )
